import io
import logging
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from flask import Flask, flash, redirect, render_template, request, send_file, session, url_for
from flask_sqlalchemy import SQLAlchemy

from extract import extract_from_pdf

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'akulab-dev-key-change-in-prod')

db_url = os.environ.get('DATABASE_URL', '')
if not db_url:
    logger.warning('DATABASE_URL není nastavena! Používám SQLite – data se smažou při každém deployi.')
    db_url = 'sqlite:///protokoly.db'
elif db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)

logger.info('Databáze: %s', 'PostgreSQL' if db_url.startswith('postgresql') else 'SQLite')

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

db = SQLAlchemy(app)


class Protocol(db.Model):
    __tablename__ = 'protocols'

    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(100), unique=True, nullable=False)
    client = db.Column(db.String(500))
    measurement_date = db.Column(db.Date)
    issue_date = db.Column(db.Date)
    auth_set = db.Column(db.String(20))
    year = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def fmt_date(self, d):
        return d.strftime('%-d. %-m. %Y') if d else '—'

    @property
    def measurement_date_str(self):
        return self.fmt_date(self.measurement_date)

    @property
    def issue_date_str(self):
        return self.fmt_date(self.issue_date)


TYPY_ZAKAZKY = ['hluková studie', 'akustická studie', 'měření hluku', 'měření vibrací', 'realizace']
STAVY_ZAKAZKY = ['založeno', 'čekám na podklady', 'zpracovává se', 'ke kontrole']


class Zakazka(db.Model):
    __tablename__ = 'zakazky'

    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(20), unique=True, nullable=False)
    typ = db.Column(db.String(50))
    zodpovedna_osoba = db.Column(db.String(100))
    stav = db.Column(db.String(50), default='založeno')
    datum_mereni = db.Column(db.Date)
    dokonceno = db.Column(db.Boolean, default=False, nullable=False)
    poznamka = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    @property
    def datum_mereni_str(self):
        return self.datum_mereni.strftime('%-d. %-m. %Y') if self.datum_mereni else '—'


with app.app_context():
    db.create_all()
    from sqlalchemy import text
    with db.engine.connect() as conn:
        conn.execute(text('ALTER TABLE zakazky ADD COLUMN IF NOT EXISTS poznamka TEXT'))
        conn.commit()


APP_PASSWORD = os.environ.get('APP_PASSWORD', '')


@app.before_request
def require_password():
    if not APP_PASSWORD:
        return
    if request.endpoint in ('login', 'static'):
        return
    if not session.get('auth'):
        return redirect(url_for('login', next=request.path))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form.get('password') == APP_PASSWORD:
            session['auth'] = True
            return redirect(request.args.get('next') or url_for('index'))
        flash('Nesprávné heslo.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


def _apply_filters(query):
    rok = request.args.get('rok', '').strip()
    auth_set = request.args.get('set', '').strip()
    search = request.args.get('search', '').strip()
    if rok:
        query = query.filter(Protocol.year == int(rok))
    if auth_set:
        query = query.filter(Protocol.auth_set == auth_set)
    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(Protocol.number.ilike(like), Protocol.client.ilike(like))
        )
    return query


@app.route('/')
def index():
    query = _apply_filters(Protocol.query)
    protocols = query.order_by(Protocol.issue_date.desc()).all()

    all_years = [r[0] for r in db.session.query(Protocol.year)
                 .distinct().order_by(Protocol.year.desc()).all() if r[0]]
    all_sets = [r[0] for r in db.session.query(Protocol.auth_set)
                .distinct().order_by(Protocol.auth_set).all() if r[0]]

    return render_template('index.html',
        protocols=protocols,
        all_years=all_years,
        all_sets=all_sets,
        current_rok=request.args.get('rok', ''),
        current_set=request.args.get('set', ''),
        current_search=request.args.get('search', ''),
        total=Protocol.query.count(),
    )


@app.route('/upload', methods=['POST'])
def upload():
    file = request.files.get('pdf')
    if not file or not file.filename.lower().endswith('.pdf'):
        flash('Vyber soubor ve formátu PDF.', 'danger')
        return redirect(url_for('index'))

    try:
        data = extract_from_pdf(file.stream)
    except Exception as e:
        flash(f'Chyba při čtení PDF: {e}', 'danger')
        return redirect(url_for('index'))

    if not data.get('number'):
        flash('Nepodařilo se načíst číslo protokolu. Zkontroluj formát souboru.', 'danger')
        return redirect(url_for('index'))

    if Protocol.query.filter_by(number=data['number']).first():
        flash(f'Protokol {data["number"]} již v databázi existuje.', 'warning')
        return redirect(url_for('index'))

    p = Protocol(
        number=data['number'],
        client=data.get('client'),
        measurement_date=data.get('measurement_date'),
        issue_date=data.get('issue_date'),
        auth_set=data.get('auth_set'),
        year=data['issue_date'].year if data.get('issue_date') else None,
    )
    db.session.add(p)
    db.session.commit()
    flash(f'Protokol {p.number} byl úspěšně přidán.', 'success')
    return redirect(url_for('index'))


@app.route('/delete/<int:id>', methods=['POST'])
def delete(id):
    p = db.get_or_404(Protocol, id)
    number = p.number
    db.session.delete(p)
    db.session.commit()
    flash(f'Protokol {number} byl smazán.', 'success')
    return redirect(url_for('index'))


@app.route('/export')
def export():
    query = _apply_filters(Protocol.query)
    protocols = query.order_by(Protocol.issue_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Protokoly'

    HDR_FILL = PatternFill('solid', fgColor='1F4E79')
    HDR_FONT = Font(color='FFFFFF', bold=True)
    ALT_FILL = PatternFill('solid', fgColor='D9E2F3')
    CENTER = Alignment(horizontal='center', vertical='center')

    headers = ['Číslo protokolu', 'Objednatel', 'Datum měření', 'Datum vydání', 'Set', 'Rok']
    widths = [24, 45, 16, 16, 8, 8]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        ws.column_dimensions[cell.column_letter].width = w

    for row, p in enumerate(protocols, 2):
        values = [p.number, p.client or '—', p.measurement_date_str,
                  p.issue_date_str, p.auth_set or '—', p.year or '—']
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.alignment = Alignment(vertical='center')
            if row % 2 == 0:
                cell.fill = ALT_FILL

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{len(protocols) + 1}'

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    rok = request.args.get('rok', '')
    s = request.args.get('set', '')
    fname = f'protokoly{"_" + rok if rok else ""}{"_" + s if s else ""}.xlsx'
    return send_file(stream, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/backup')
def backup():
    protocols = Protocol.query.order_by(Protocol.issue_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Protokoly'

    HDR_FILL = PatternFill('solid', fgColor='1F4E79')
    HDR_FONT = Font(color='FFFFFF', bold=True)
    ALT_FILL = PatternFill('solid', fgColor='D9E2F3')
    CENTER = Alignment(horizontal='center', vertical='center')

    headers = ['Číslo protokolu', 'Objednatel', 'Datum měření', 'Datum vydání', 'Set', 'Rok']
    widths = [24, 45, 16, 16, 8, 8]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        ws.column_dimensions[cell.column_letter].width = w

    for row, p in enumerate(protocols, 2):
        values = [p.number, p.client or '', p.measurement_date_str,
                  p.issue_date_str, p.auth_set or '', p.year or '']
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.alignment = Alignment(vertical='center')
            if row % 2 == 0:
                cell.fill = ALT_FILL

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{len(protocols) + 1}'

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    from datetime import date as date_cls
    fname = f'zaloha_protokoly_{date_cls.today().isoformat()}.xlsx'
    return send_file(stream, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/restore', methods=['POST'])
def restore():
    file = request.files.get('xlsx')
    if not file or not file.filename.lower().endswith('.xlsx'):
        flash('Vyber soubor ve formátu .xlsx (záloha protokolů).', 'danger')
        return redirect(url_for('index'))

    try:
        wb = openpyxl.load_workbook(file.stream, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'Chyba při čtení souboru: {e}', 'danger')
        return redirect(url_for('index'))

    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        number = str(row[0]).strip()
        if not number:
            continue

        if Protocol.query.filter_by(number=number).first():
            skipped += 1
            continue

        def parse_date(val):
            if not val or str(val).strip() in ('', '—'):
                return None
            import re
            parts = re.findall(r'\d+', str(val))
            if len(parts) == 3:
                try:
                    return __import__('datetime').date(int(parts[2]), int(parts[1]), int(parts[0]))
                except ValueError:
                    return None
            return None

        client = str(row[1]).strip() if row[1] else None
        measurement_date = parse_date(row[2])
        issue_date = parse_date(row[3])
        auth_set = str(row[4]).strip() if row[4] else None
        year = issue_date.year if issue_date else None

        p = Protocol(number=number, client=client,
                     measurement_date=measurement_date, issue_date=issue_date,
                     auth_set=auth_set, year=year)
        db.session.add(p)
        imported += 1

    db.session.commit()
    flash(f'Obnoveno: {imported} protokolů nahráno, {skipped} přeskočeno (již existují).', 'success')
    return redirect(url_for('index'))


@app.route('/zakazky')
def zakazky():
    view = request.args.get('view', 'aktivni')
    historie = (view == 'historie')

    query = Zakazka.query.filter_by(dokonceno=historie)

    typ = request.args.get('typ', '').strip()
    osoba = request.args.get('osoba', '').strip()
    stav = request.args.get('stav', '').strip()
    if typ:
        query = query.filter(Zakazka.typ == typ)
    if osoba:
        query = query.filter(Zakazka.zodpovedna_osoba.ilike(f'%{osoba}%'))
    if stav:
        query = query.filter(Zakazka.stav == stav)

    zakazky_list = query.order_by(Zakazka.created_at.desc()).all()

    all_osoby = [r[0] for r in db.session.query(Zakazka.zodpovedna_osoba)
                 .distinct().order_by(Zakazka.zodpovedna_osoba).all() if r[0]]

    return render_template('zakazky.html',
        zakazky=zakazky_list,
        historie=historie,
        view=view,
        typy=TYPY_ZAKAZKY,
        stavy=STAVY_ZAKAZKY,
        all_osoby=all_osoby,
        current_typ=typ,
        current_osoba=osoba,
        current_stav=stav,
        aktivni_count=Zakazka.query.filter_by(dokonceno=False).count(),
        historie_count=Zakazka.query.filter_by(dokonceno=True).count(),
    )


@app.route('/zakazky/new', methods=['POST'])
def zakazka_new():
    number = request.form.get('number', '').strip()
    if not number:
        flash('Číslo zakázky je povinné.', 'danger')
        return redirect(url_for('zakazky'))
    if Zakazka.query.filter_by(number=number).first():
        flash(f'Zakázka {number} již existuje.', 'warning')
        return redirect(url_for('zakazky'))

    datum_raw = request.form.get('datum_mereni', '').strip()
    datum = None
    if datum_raw:
        try:
            from datetime import date as date_cls
            datum = date_cls.fromisoformat(datum_raw)
        except ValueError:
            pass

    z = Zakazka(
        number=number,
        typ=request.form.get('typ') or None,
        zodpovedna_osoba=request.form.get('zodpovedna_osoba', '').strip() or None,
        stav=request.form.get('stav', 'založeno'),
        datum_mereni=datum,
        poznamka=request.form.get('poznamka', '').strip() or None,
    )
    db.session.add(z)
    db.session.commit()
    flash(f'Zakázka {number} byla přidána.', 'success')
    return redirect(url_for('zakazky'))


@app.route('/zakazky/<int:id>/edit', methods=['POST'])
def zakazka_edit(id):
    z = db.get_or_404(Zakazka, id)

    datum_raw = request.form.get('datum_mereni', '').strip()
    datum = None
    if datum_raw:
        try:
            from datetime import date as date_cls
            datum = date_cls.fromisoformat(datum_raw)
        except ValueError:
            pass

    z.typ = request.form.get('typ') or None
    z.zodpovedna_osoba = request.form.get('zodpovedna_osoba', '').strip() or None
    z.stav = request.form.get('stav', z.stav)
    z.datum_mereni = datum
    z.poznamka = request.form.get('poznamka', '').strip() or None
    db.session.commit()
    flash(f'Zakázka {z.number} byla upravena.', 'success')
    return redirect(url_for('zakazky', view='historie' if z.dokonceno else 'aktivni'))


@app.route('/zakazky/<int:id>/done', methods=['POST'])
def zakazka_done(id):
    z = db.get_or_404(Zakazka, id)
    z.dokonceno = not z.dokonceno
    db.session.commit()
    if z.dokonceno:
        flash(f'Zakázka {z.number} byla dokončena.', 'success')
        return redirect(url_for('zakazky'))
    else:
        flash(f'Zakázka {z.number} byla vrácena mezi aktivní.', 'success')
        return redirect(url_for('zakazky', view='historie'))


@app.route('/zakazky/<int:id>/delete', methods=['POST'])
def zakazka_delete(id):
    z = db.get_or_404(Zakazka, id)
    number = z.number
    was_done = z.dokonceno
    db.session.delete(z)
    db.session.commit()
    flash(f'Zakázka {number} byla smazána.', 'success')
    return redirect(url_for('zakazky', view='historie' if was_done else 'aktivni'))


if __name__ == '__main__':
    app.run(debug=True)
